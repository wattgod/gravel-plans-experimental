#!/usr/bin/env python3
"""
Extract all exercises and URLs from PN library Excel file
Creates a simple CSV mapping file for easier use
"""

import openpyxl
import re
import csv
import sys

def extract_all_exercises(excel_path):
    """Extract all exercise name -> URL mappings from PN library"""
    wb = openpyxl.load_workbook(excel_path)
    
    url_pattern = re.compile(r'https?://[^\s,"]+')
    exercises = {}
    
    print("Extracting exercises from all sheets...")
    print("="*70)
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        if 'Contents' in sheet_name:
            continue  # Skip table of contents
        
        ws = wb[sheet_name]
        sheet_count = 0
        
        # Check every cell
        for row in ws.iter_rows():
            for cell in row:
                url = None
                name = None
                
                # Method 1: Check hyperlink
                if cell.hyperlink and cell.hyperlink.target:
                    url = cell.hyperlink.target
                    if cell.value:
                        name = str(cell.value).strip()
                
                # Method 2: Check cell value for embedded URL
                if not url and cell.value:
                    val = str(cell.value)
                    if 'vimeo.com' in val or ('http' in val and len(val) > 20):
                        url_match = url_pattern.search(val)
                        if url_match:
                            url = url_match.group(0)
                            # Extract name from cell
                            name = re.sub(url_pattern, '', val).strip()
                            name = re.sub(r'^[,"\']+', '', name)
                            name = re.sub(r'["\']+.*$', '', name)
                
                # Method 3: If we have URL but no name, check adjacent cells
                if url and (not name or len(name) < 3):
                    # Check cell to the right
                    right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if right_cell.value:
                        potential_name = str(right_cell.value).strip()
                        if len(potential_name) > 3 and 'http' not in potential_name.lower():
                            name = potential_name
                    
                    # Check cell above (might be category header)
                    if not name and cell.row > 1:
                        above_cell = ws.cell(row=cell.row - 1, column=cell.column)
                        if above_cell.value:
                            potential_name = str(above_cell.value).strip()
                            if len(potential_name) > 3 and 'http' not in potential_name.lower():
                                name = potential_name
                
                # Clean and store
                if url and name:
                    name = name.strip()
                    # Remove common prefixes that aren't part of exercise name
                    name = re.sub(r'^(Bodyweight|Dumbbell|KB|Kettlebell|Band|TRX|Cable|Barbell)\s+', '', name, flags=re.IGNORECASE)
                    name = re.sub(r'\s+', ' ', name)  # Normalize whitespace
                    
                    # Filter out non-exercise names
                    skip_words = ['instructions', 'exercise', 'bodyweight', 'band', 'dumbbell', 
                                 'category', 'type', 'movement', 'sheet', 'index']
                    if (len(name) > 3 and 
                        not any(skip in name.lower() for skip in skip_words) and
                        name[0].isupper()):  # Exercise names usually start with capital
                        
                        # Store (keep first occurrence if duplicate)
                        if name not in exercises:
                            exercises[name] = url
                            sheet_count += 1
        
        if sheet_count > 0:
            print(f"{sheet_name}: {sheet_count} exercises")
    
    print(f"\nTotal unique exercises: {len(exercises)}")
    return exercises

def save_to_csv(exercises, output_path):
    """Save exercises to CSV file"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Exercise Name', 'Video URL'])
        for name, url in sorted(exercises.items()):
            writer.writerow([name, url])
    print(f"\nSaved to: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_pn_exercises.py <excel_file> [output_csv]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else excel_path.replace('.xlsx', '_exercises.csv')
    
    exercises = extract_all_exercises(excel_path)
    save_to_csv(exercises, output_path)
    
    print(f"\nSample exercises (first 10):")
    for i, (name, url) in enumerate(sorted(exercises.items())[:10]):
        print(f"  {i+1}. {name}")

