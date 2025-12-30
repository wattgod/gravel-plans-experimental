#!/usr/bin/env python3
"""
Update Strength Templates to Use Precision Nutrition Video Library
Replaces YouTube URLs with PN library URLs
"""

import sys
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import csv

def load_pn_library(library_path: str, manual_mapping_path: str = None) -> Dict[str, str]:
    """
    Load PN exercise library from CSV or Excel
    
    Args:
        library_path: Path to PN library CSV/Excel
        manual_mapping_path: Optional path to manual mapping CSV
    
    Returns:
        Dict mapping exercise names to video URLs
    """
    pn_library = {}
    
    # Load manual mappings first (these take precedence)
    if manual_mapping_path and Path(manual_mapping_path).exists():
        with open(manual_mapping_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                template_name = row.get('Exercise Name', '').strip()
                pn_name = row.get('PN Exercise Name', '').strip()
                url = row.get('PN Video URL', '').strip()
                
                if template_name and url and url != 'NOT_FOUND':
                    # Add both original and lowercase versions
                    pn_library[template_name] = url
                    pn_library[template_name.lower()] = url
                    # Also add normalized version (Push-Up -> Pushup)
                    normalized = normalize_exercise_name(template_name)
                    pn_library[normalized] = url
                    pn_library[normalized.lower()] = url
                    if pn_name:
                        pn_library[pn_name.lower()] = url
    
    if library_path.endswith('.csv'):
        with open(library_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Try common column names
                ex_name = (row.get('Exercise Name') or 
                          row.get('Exercise') or 
                          row.get('Name') or
                          row.get('Exercise Name (Men)') or
                          row.get('Exercise Name (Women)'))
                
                url = (row.get('Video URL') or 
                      row.get('URL') or 
                      row.get('Link') or
                      row.get('Video Link') or
                      row.get('Men\'s Video URL') or
                      row.get('Women\'s Video URL'))
                
                if ex_name and url:
                    # Normalize exercise name
                    ex_name_clean = ex_name.strip()
                    pn_library[ex_name_clean] = url.strip()
                    
                    # Also add lowercase version for matching
                    pn_library[ex_name_clean.lower()] = url.strip()
                    
                    # Add variations (Push-Up -> Pushup)
                    if 'pushup' in ex_name_clean.lower():
                        pn_library[ex_name_clean.lower().replace('pushup', 'push-up')] = url.strip()
                    if 'push-up' in ex_name_clean.lower():
                        pn_library[ex_name_clean.lower().replace('push-up', 'pushup')] = url.strip()
    
    elif library_path.endswith(('.xlsx', '.xls')):
        try:
            import pandas as pd
            import openpyxl
        except ImportError:
            print("ERROR: pandas/openpyxl not installed. Install with: pip install pandas openpyxl")
            sys.exit(1)
        
        try:
            # Use openpyxl to read hyperlinks and cell values
            wb = openpyxl.load_workbook(library_path, data_only=True)
            
            # Check all sheets, but prioritize index sheets
            sheets_to_check = []
            for sheet_name in wb.sheetnames:
                if 'index' in sheet_name.lower() and 'alphabetical' in sheet_name.lower():
                    sheets_to_check.insert(0, sheet_name)  # Prioritize index sheets
                elif 'index' in sheet_name.lower():
                    sheets_to_check.append(sheet_name)
                else:
                    sheets_to_check.append(sheet_name)
            
            print(f"   Checking {len(sheets_to_check)} sheets for exercises...")
            
            url_pattern = re.compile(r'https?://[^\s,"]+')
            
            for sheet_name in sheets_to_check:
                ws = wb[sheet_name]
                sheet_count = 0
                
                for row in ws.iter_rows():
                    for cell in row:
                        if not cell.value:
                            continue
                        
                        cell_value = str(cell.value)
                        
                        # Check for hyperlink
                        url = None
                        if cell.hyperlink and cell.hyperlink.target:
                            url = cell.hyperlink.target
                        elif url_pattern.search(cell_value):
                            # Extract URL from cell text (format: "https://vimeo.com/123","Exercise Name")
                            match = url_pattern.search(cell_value)
                            if match:
                                url = match.group(0)
                        
                        if url:
                            # Extract exercise name
                            # Format might be: "https://vimeo.com/123","Exercise Name" or just "Exercise Name" with hyperlink
                            ex_name = None
                            
                            # If URL is in text, extract name after it
                            if url in cell_value:
                                parts = cell_value.split(url)
                                if len(parts) > 1:
                                    # Name is after URL, might be in quotes
                                    name_part = parts[1].strip()
                                    name_part = re.sub(r'^[,"\']+', '', name_part)  # Remove leading quotes/commas
                                    name_part = re.sub(r'["\']+.*$', '', name_part)  # Remove trailing quotes and rest
                                    if name_part and len(name_part) > 2:
                                        ex_name = name_part.strip()
                            
                            # If no name extracted, use cell value (might be just the name with hyperlink)
                            if not ex_name:
                                # Remove URL from cell value
                                ex_name = re.sub(url_pattern, '', cell_value).strip()
                                ex_name = re.sub(r'^[,"\']+', '', ex_name)
                                ex_name = re.sub(r'["\']+.*$', '', ex_name)
                            
                            # Clean up exercise name
                            if ex_name:
                                ex_name = ex_name.strip()
                                # Remove common prefixes/suffixes
                                ex_name = re.sub(r'^(Bodyweight|Dumbbell|KB|Kettlebell|Band|TRX|Cable)\s+', '', ex_name, flags=re.IGNORECASE)
                                
                                if ex_name and len(ex_name) > 2 and ex_name.lower() != 'nan':
                                    pn_library[ex_name] = url
                                    pn_library[ex_name.lower()] = url
                                    sheet_count += 1
                
                if sheet_count > 0:
                    print(f"   ✓ Found {sheet_count} exercises in '{sheet_name}'")
            
            print(f"   Total unique exercises loaded: {len(set(pn_library.values()))}")
            
        except Exception as e:
            print(f"ERROR reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    return pn_library

def normalize_exercise_name(ex_name: str) -> str:
    """Normalize exercise name for matching"""
    # Remove parenthetical notes
    normalized = re.sub(r'\s*\([^)]+\)', '', ex_name)
    # Remove variation markers
    normalized = re.sub(r'\s*w/\s*[^─\n]+', '', normalized)
    normalized = re.sub(r'\s*\(hands[^)]*\)', '', normalized)
    # Remove set/rep markers like "A1", "B2", etc.
    normalized = re.sub(r'^[A-Z]\d+\s+', '', normalized)
    # Normalize hyphens and spacing
    normalized = normalized.replace('Push-Up', 'Pushup')
    normalized = normalized.replace('Push-up', 'Pushup')
    normalized = normalized.replace('push-up', 'pushup')
    # Clean whitespace
    normalized = ' '.join(normalized.split())
    return normalized.strip()

def find_pn_match(exercise_name: str, pn_library: Dict[str, str]) -> Tuple[Optional[str], str]:
    """
    Find matching exercise in PN library
    
    Returns:
        (PN URL if found, match method used)
    """
    # First check original name (for manual mappings)
    if exercise_name in pn_library:
        return pn_library[exercise_name], "direct (original)"
    
    if exercise_name.lower() in pn_library:
        return pn_library[exercise_name.lower()], "lowercase (original)"
    
    normalized = normalize_exercise_name(exercise_name)
    
    # Direct match
    if normalized in pn_library:
        return pn_library[normalized], "direct"
    
    # Lowercase match
    if normalized.lower() in pn_library:
        return pn_library[normalized.lower()], "lowercase"
    
    # Partial match - check if any PN exercise contains key words
    normalized_words = set(normalized.lower().split())
    best_match = None
    best_score = 0
    
    for pn_exercise, url in pn_library.items():
        if not url or url == 'nan':
            continue
        
        pn_words = set(pn_exercise.lower().split())
        # Count matching words
        matches = len(normalized_words & pn_words)
        
        # Prefer matches with 2+ words
        if matches >= 2 and matches > best_score:
            best_match = url
            best_score = matches
    
    if best_match:
        return best_match, f"partial ({best_score} words)"
    
    return None, "no match"

def update_template_urls(template_text: str, pn_library: Dict[str, str]) -> Tuple[str, List[str]]:
    """
    Replace YouTube URLs with PN library URLs
    
    Returns:
        (updated template text, list of unmatched exercises)
    """
    # Pattern: Exercise name followed by URL on next line
    pattern = r'([A-Z][^→\n]+?)\s*─\s*[^\n]*\n\s*→\s*(https://www\.youtube\.com/watch\?v=[^\s\n]+)'
    
    unmatched = []
    
    def replace_match(match):
        exercise_line = match.group(1).strip()
        youtube_url = match.group(2)
        
        # Extract exercise name
        exercise_name = re.sub(r'\s*\([^)]+\)', '', exercise_line).strip()
        # Remove set markers
        exercise_name = re.sub(r'^[A-Z]\d+\s+', '', exercise_name).strip()
        
        # Find PN match
        pn_url, match_method = find_pn_match(exercise_name, pn_library)
        
        if pn_url:
            # Replace with PN URL
            return f"{exercise_line} ─ [same reps/sets]\n     → {pn_url}"
        else:
            unmatched.append(f"{exercise_name} (was: {youtube_url})")
            # Keep YouTube URL
            return match.group(0)
    
    updated_text = re.sub(pattern, replace_match, template_text)
    
    return updated_text, unmatched

def main():
    if len(sys.argv) < 3:
        print("Usage: python update_templates_to_pn.py <pn_library_file> <template_file> [output_file]")
        print("\nExample:")
        print("  python update_templates_to_pn.py pn_library.csv MASTER_TEMPLATES_V2.md updated_templates.md")
        sys.exit(1)
    
    library_path = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else template_path.replace('.md', '_pn_updated.md')
    
    print("=" * 70)
    print("PN Exercise Library Template Updater")
    print("=" * 70)
    
    # Load PN library
    print(f"\n📥 Loading PN library from: {library_path}")
    manual_mapping = library_path.replace('.csv', '_manual_mapping.csv').replace('.xlsx', '_manual_mapping.csv')
    pn_library = load_pn_library(library_path, manual_mapping_path=manual_mapping)
    print(f"   ✓ Loaded {len(pn_library)} exercises from PN library")
    
    # Load templates
    print(f"\n📥 Loading templates from: {template_path}")
    with open(template_path, 'r', encoding='utf-8') as f:
        template_content = f.read()
    
    # Update URLs
    print(f"\n🔄 Updating YouTube URLs to PN library URLs...")
    updated_content, unmatched = update_template_urls(template_content, pn_library)
    
    # Save updated templates
    print(f"\n💾 Saving updated templates to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(updated_content)
    
    # Report results
    print(f"\n✅ Update complete!")
    print(f"\n📊 Results:")
    
    # Count replacements
    youtube_count = len(re.findall(r'https://www\.youtube\.com/watch\?v=', updated_content))
    pn_count = len(re.findall(r'→\s+(?!https://www\.youtube\.com)', updated_content))
    
    print(f"   • YouTube URLs remaining: {youtube_count}")
    print(f"   • PN URLs added: {pn_count}")
    print(f"   • Unmatched exercises: {len(unmatched)}")
    
    if unmatched:
        print(f"\n⚠️  Unmatched exercises (still using YouTube):")
        for ex in unmatched[:10]:  # Show first 10
            print(f"   - {ex}")
        if len(unmatched) > 10:
            print(f"   ... and {len(unmatched) - 10} more")
    
    print(f"\n📝 Next steps:")
    print(f"   1. Review updated templates: {output_path}")
    print(f"   2. Manually match any unmatched exercises")
    print(f"   3. Replace MASTER_TEMPLATES_V2.md with updated version")
    print(f"   4. Regenerate ZWO files")

if __name__ == "__main__":
    main()

