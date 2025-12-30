#!/usr/bin/env python3
"""
Extract exercise names and URLs from PN Excel library
Handles hyperlinks and matches them with exercise names
"""

import openpyxl
import csv
import re
import sys
from pathlib import Path

def extract_exercises_from_excel(excel_path):
    """Extract all exercise name -> URL mappings from Excel"""
    # Load WITHOUT data_only to preserve formulas
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    exercises = {}
    hyperlink_pattern = re.compile(r'=HYPERLINK\("([^"]+)","([^"]+)"\)', re.IGNORECASE)
    
    print("Extracting exercises from Excel file (reading HYPERLINK formulas)...")
    print("="*70)
    
    # Process all sheets except table of contents
    for sheet_name in wb.sheetnames:
        if 'Contents' in sheet_name or 'Tutorial' in sheet_name or 'FAQ' in sheet_name:
            continue
        
        ws = wb[sheet_name]
        sheet_count = 0
        
        # Check every cell for HYPERLINK formulas
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                
                cell_value = str(cell.value)
                
                # Check for HYPERLINK formula
                if 'HYPERLINK' in cell_value.upper() and cell_value.startswith('='):
                    match = hyperlink_pattern.search(cell_value)
                    if match:
                        url = match.group(1)
                        name = match.group(2).strip()
                        
                        # Clean name
                        name = re.sub(r'\s+', ' ', name)
                        
                        # Filter out non-exercise names
                        skip_words = ['instructions', 'exercise', 'category', 'type', 'movement', 'sheet', 'index', 
                                     'demos', 'craig', 'jen', 'horizontal', 'vertical', 'pressing', 'pulling',
                                     'table of contents', 'tutorial']
                        if (len(name) > 3 and 
                            not any(skip in name.lower() for skip in skip_words) and
                            name[0].isupper() and
                            name not in ['Barbell', 'Dumbbell', 'Pushup', 'Cable or band', 'Band', 'Cable']):
                            
                            # Store (prefer first occurrence)
                            if name not in exercises:
                                exercises[name] = url
                                sheet_count += 1
                                if sheet_count <= 3:
                                    print(f"  ✓ {name[:50]:50s} -> {url}")
        
        if sheet_count > 0:
            print(f"  {sheet_name}: {sheet_count} exercises")
    
    print(f"\n{'='*70}")
    print(f"Total unique exercises extracted: {len(exercises)}")
    return exercises

def save_to_csv(exercises, output_path):
    """Save exercises to CSV"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Exercise Name', 'Video URL'])
        for name, url in sorted(exercises.items()):
            writer.writerow([name, url])
    print(f"\nSaved {len(exercises)} exercises to: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_pn_from_excel.py <excel_file> [output_csv]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else excel_path.replace('.xlsx', '_exercises.csv')
    
    exercises = extract_exercises_from_excel(excel_path)
    save_to_csv(exercises, output_path)
    
    print(f"\nSample exercises (first 15):")
    for i, (name, url) in enumerate(sorted(exercises.items())[:15]):
        print(f"  {i+1:2d}. {name:40s} -> {url}")

